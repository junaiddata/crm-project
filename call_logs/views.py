import json

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from crm_project.alabama_db import tpl, pick, is_alabama
from .models import (
    CallLog, CallLead, AlabamaCallLog, AlabamaCallLead,
    ExcludedNumber, AlabamaExcludedNumber,
)
from .serializers import (
    CallLogCreateSerializer, CallLogSerializer,
    AlabamaCallLogCreateSerializer, AlabamaCallLogSerializer,
)


# ── REST API ─────────────────────────────────────────────────────────────────

class LogCallView(APIView):
    """POST /api/calls/log/ — called by the Android app."""
    authentication_classes = []
    permission_classes     = []

    def post(self, request):
        CreateSer = pick(request, CallLogCreateSerializer, AlabamaCallLogCreateSerializer)
        CallLeadM = pick(request, CallLead, AlabamaCallLead)
        serializer = CreateSer(data=request.data)
        if serializer.is_valid():
            log = serializer.save()

            if log.direction == 'incoming':
                # Auto-create a CallLead for every incoming call
                CallLeadM.objects.get_or_create(
                    call_log=log,
                    defaults=dict(
                        caller_number=log.caller_number,
                        call_status=log.status,
                        duration=log.duration,
                        call_time=log.timestamp,
                        received_by=log.received_by,
                        sim=log.sim,
                    )
                )

            elif log.direction == 'outgoing' and log.status == 'answered':
                # Check if we called back a missed lead within the last 24 hours
                cutoff = log.timestamp - timezone.timedelta(hours=24)
                missed_leads = CallLeadM.objects.filter(
                    caller_number=log.caller_number,
                    call_status='missed',
                    return_called=False,
                    call_time__gte=cutoff,
                    call_time__lte=log.timestamp,
                )
                if missed_leads.exists():
                    missed_leads.update(
                        return_called=True,
                        return_called_at=log.timestamp,
                    )

            return Response({
                'id':            log.id,
                'caller_number': log.caller_number,
                'status':        log.status,
                'duration':      log.duration,
                'timestamp':     str(log.timestamp),
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CallLogListView(APIView):
    """GET /api/calls/ — list with optional filters."""

    def get(self, request):
        CallLogM = pick(request, CallLog, AlabamaCallLog)
        ListSer = pick(request, CallLogSerializer, AlabamaCallLogSerializer)
        qs = CallLogM.objects.all()

        status_f  = request.GET.get('status', '').strip()
        device_f  = request.GET.get('received_by', '').strip()
        sim_f     = request.GET.get('sim', '').strip()
        from_date = request.GET.get('from_date', '').strip()
        to_date   = request.GET.get('to_date', '').strip()

        if status_f:   qs = qs.filter(status=status_f)
        if device_f:   qs = qs.filter(received_by=device_f)
        if sim_f:      qs = qs.filter(sim=sim_f)
        if from_date:  qs = qs.filter(timestamp__date__gte=from_date)
        if to_date:    qs = qs.filter(timestamp__date__lte=to_date)

        return Response(ListSer(qs, many=True).data)


# ── Calls dashboard ───────────────────────────────────────────────────────────

def excluded_numbers(request):
    """List of numbers to hide from Call Log / Call Leads (per site)."""
    ExcludedM = pick(request, ExcludedNumber, AlabamaExcludedNumber)
    return list(ExcludedM.objects.values_list('number', flat=True))


def _filtered_call_logs(request):
    """Call Log rows after exclusions + the status/device/SIM/date filters —
    the exact set shown on the Call Log dashboard (and its Excel export)."""
    CallLogM = pick(request, CallLog, AlabamaCallLog)
    qs = CallLogM.objects.exclude(caller_number__in=excluded_numbers(request))
    status_f  = request.GET.get('status', '').strip()
    device_f  = request.GET.get('device', '').strip()
    sim_f     = request.GET.get('sim', '').strip()
    from_date = request.GET.get('from_date', '').strip()
    to_date   = request.GET.get('to_date', '').strip()
    if status_f:   qs = qs.filter(status=status_f)
    if device_f:   qs = qs.filter(received_by=device_f)
    if sim_f:      qs = qs.filter(sim=sim_f)
    if from_date:  qs = qs.filter(timestamp__date__gte=from_date)
    if to_date:    qs = qs.filter(timestamp__date__lte=to_date)
    return qs


@login_required
def calls_dashboard(request):
    CallLogM = pick(request, CallLog, AlabamaCallLog)
    today = timezone.now().date()

    excluded = excluded_numbers(request)

    qs = _filtered_call_logs(request)

    status_f  = request.GET.get('status', '').strip()
    device_f  = request.GET.get('device', '').strip()
    sim_f     = request.GET.get('sim', '').strip()
    from_date = request.GET.get('from_date', '').strip()
    to_date   = request.GET.get('to_date', '').strip()

    # Stats honor the device / SIM / date filters so the cards match what is
    # being viewed. The status filter is intentionally ignored here because the
    # cards already break the results down by status. With no scoping filter
    # active they fall back to today's numbers.
    stats_qs = CallLogM.objects.exclude(caller_number__in=excluded)
    if device_f:   stats_qs = stats_qs.filter(received_by=device_f)
    if sim_f:      stats_qs = stats_qs.filter(sim=sim_f)
    if from_date:  stats_qs = stats_qs.filter(timestamp__date__gte=from_date)
    if to_date:    stats_qs = stats_qs.filter(timestamp__date__lte=to_date)

    stats_scoped = bool(device_f or sim_f or from_date or to_date)
    if not stats_scoped:
        stats_qs = stats_qs.filter(timestamp__date=today)

    stat_total    = stats_qs.count()
    stat_answered = stats_qs.filter(status='answered').count()
    stat_missed   = stats_qs.filter(status='missed').count()

    all_devices = CallLogM.objects.values_list('received_by', flat=True).distinct().order_by('received_by')
    all_sims    = (CallLogM.objects.exclude(sim='')
                   .values_list('sim', flat=True).distinct().order_by('sim'))

    return render(request, tpl(request, 'calls_dashboard.html'), {
        'calls':          qs,
        'shown_count':    qs.count(),
        'stat_total':     stat_total,
        'stat_answered':  stat_answered,
        'stat_missed':    stat_missed,
        'stats_scoped':   stats_scoped,
        'all_devices':    all_devices,
        'all_sims':       all_sims,
        'active_status':  status_f,
        'active_device':  device_f,
        'active_sim':     sim_f,
        'from_date':      from_date,
        'to_date':        to_date,
    })


# ── Call Leads dashboard ──────────────────────────────────────────────────────

def _filtered_call_leads(request):
    """Call Lead rows after exclusions + tab + shared filters — the exact set
    shown on the Call Leads dashboard (and its Excel export)."""
    CallLeadM = pick(request, CallLead, AlabamaCallLead)
    tab       = request.GET.get('tab', 'active')
    from_date = request.GET.get('from_date', '').strip()
    to_date   = request.GET.get('to_date', '').strip()
    search    = request.GET.get('q', '').strip()
    sim_f     = request.GET.get('sim', '').strip()
    device_f  = request.GET.get('device', '').strip()

    qs = CallLeadM.objects.select_related('call_log').exclude(
        caller_number__in=excluded_numbers(request))

    if tab == 'junk':
        qs = qs.filter(lead_status__in=['junk', 'irrelevant'])
    elif tab == 'missed':
        qs = qs.filter(lead_status='active', call_status='missed')
    elif tab == 'followup':
        qs = qs.filter(lead_status='active', follow_up__isnull=False)
    else:
        qs = qs.filter(lead_status='active')

    if from_date:  qs = qs.filter(call_time__date__gte=from_date)
    if to_date:    qs = qs.filter(call_time__date__lte=to_date)
    if search:     qs = qs.filter(caller_number__icontains=search)
    if sim_f:      qs = qs.filter(sim=sim_f)
    if device_f:   qs = qs.filter(received_by=device_f)
    return qs


@login_required
def call_leads_dashboard(request):
    CallLeadM = pick(request, CallLead, AlabamaCallLead)
    tab       = request.GET.get('tab', 'active')
    from_date = request.GET.get('from_date', '').strip()
    to_date   = request.GET.get('to_date', '').strip()
    search    = request.GET.get('q', '').strip()
    sim_f     = request.GET.get('sim', '').strip()
    device_f  = request.GET.get('device', '').strip()

    excluded = excluded_numbers(request)
    qs = _filtered_call_leads(request)

    # Distinct SIM lines across all leads, for the "number" filter dropdown.
    all_sims = (CallLeadM.objects.exclude(sim='')
                .values_list('sim', flat=True).distinct().order_by('sim'))
    # Distinct phones / devices that handled the calls, for the "phone" filter.
    all_devices = (CallLeadM.objects.exclude(received_by='')
                   .values_list('received_by', flat=True).distinct().order_by('received_by'))

    today = timezone.now().date()

    # Stat cards exclude hidden numbers and honor the shared filters (date range,
    # search, SIM, phone) so they match the current view. They do NOT apply the
    # tab filter — each card is its own category.
    counts_qs = CallLeadM.objects.exclude(caller_number__in=excluded)
    if from_date:  counts_qs = counts_qs.filter(call_time__date__gte=from_date)
    if to_date:    counts_qs = counts_qs.filter(call_time__date__lte=to_date)
    if search:     counts_qs = counts_qs.filter(caller_number__icontains=search)
    if sim_f:      counts_qs = counts_qs.filter(sim=sim_f)
    if device_f:   counts_qs = counts_qs.filter(received_by=device_f)

    return render(request, tpl(request, 'call_leads.html'), {
        'leads':        qs,
        'total_active': counts_qs.filter(lead_status='active').count(),
        'total_missed': counts_qs.filter(lead_status='active', call_status='missed').count(),
        'total_followup': counts_qs.filter(lead_status='active', follow_up__isnull=False, follow_up__lte=today).count(),
        'total_junk':   counts_qs.filter(lead_status__in=['junk','irrelevant']).count(),
        'tab':          tab,
        'from_date':    from_date,
        'to_date':      to_date,
        'search':       search,
        'all_sims':     all_sims,
        'active_sim':   sim_f,
        'all_devices':  all_devices,
        'active_device': device_f,
    })


# ── AJAX endpoints for Call Leads ─────────────────────────────────────────────

@login_required
@csrf_exempt
def update_call_lead(request, pk):
    CallLeadM = pick(request, CallLead, AlabamaCallLead)
    lead = get_object_or_404(CallLeadM, pk=pk)

    # File upload (multipart)
    if request.method == 'POST' and request.FILES.get('quotation_file'):
        f = request.FILES['quotation_file']
        if lead.quotation_file:
            lead.quotation_file.delete(save=False)
        lead.quotation_file     = f
        lead.quotation_filename = f.name
        lead.save()
        return JsonResponse({
            'ok': True,
            'url':  request.build_absolute_uri(lead.quotation_file.url),
            'name': lead.quotation_filename,
        })

    # Delete quotation file
    if request.method == 'POST' and request.POST.get('delete_quotation'):
        if lead.quotation_file:
            lead.quotation_file.delete(save=False)
        lead.quotation_file     = None
        lead.quotation_filename = ''
        lead.save()
        return JsonResponse({'ok': True})

    # JSON field update
    if request.method == 'POST':
        data = json.loads(request.body)
        if 'query'          in data: lead.query          = data['query']
        if 'follow_up_note' in data: lead.follow_up_note = data['follow_up_note']
        if 'notes'          in data: lead.notes          = data['notes']
        if 'follow_up'      in data: lead.follow_up      = data['follow_up'] or None
        lead.save()
        return JsonResponse({'ok': True})

    return JsonResponse({'ok': False}, status=405)


@login_required
@csrf_exempt
@require_POST
def toggle_return_called(request, pk):
    CallLeadM = pick(request, CallLead, AlabamaCallLead)
    lead = get_object_or_404(CallLeadM, pk=pk)
    lead.return_called = not lead.return_called
    lead.return_called_at = timezone.now() if lead.return_called else None
    lead.save()
    return JsonResponse({'ok': True, 'return_called': lead.return_called})


@login_required
@csrf_exempt
@require_POST
def set_lead_status(request, pk):
    CallLeadM = pick(request, CallLead, AlabamaCallLead)
    lead = get_object_or_404(CallLeadM, pk=pk)
    data = json.loads(request.body)
    new_status = data.get('status', 'active')
    if new_status in ['active', 'junk', 'irrelevant']:
        lead.lead_status = new_status
        lead.save()
    return JsonResponse({'ok': True, 'status': lead.lead_status})


# ── Excluded numbers ──────────────────────────────────────────────────────────

@login_required
def excluded_numbers_page(request):
    """Manage the hidden-numbers list. Calls to/from these numbers are dropped
    from the Call Log and Call Leads dashboards for this site."""
    ExcludedM = pick(request, ExcludedNumber, AlabamaExcludedNumber)
    base = '/alabama' if is_alabama(request) else ''

    if request.method == 'POST':
        # Accept one or many numbers — one per line (also split on commas).
        raw    = request.POST.get('number', '')
        note   = request.POST.get('note', '').strip()
        seen   = set()
        for line in raw.replace(',', '\n').splitlines():
            number = line.strip()
            if number and number not in seen:
                seen.add(number)
                ExcludedM.objects.get_or_create(number=number, defaults={'note': note})
        return redirect(base + '/call-exclusions/')

    return render(request, tpl(request, 'call_exclusions.html'), {
        'excluded':  ExcludedM.objects.all(),
        'base':      base,
    })


@login_required
@require_POST
def delete_excluded_number(request, pk):
    ExcludedM = pick(request, ExcludedNumber, AlabamaExcludedNumber)
    ExcludedM.objects.filter(pk=pk).delete()
    base = '/alabama' if is_alabama(request) else ''
    return redirect(base + '/call-exclusions/')


# ── Excel exports ─────────────────────────────────────────────────────────────

def _fmt_duration(secs):
    """Seconds → m:ss (e.g. 125 → '2:05'). 0/blank → '—'."""
    secs = int(secs or 0)
    if secs <= 0:
        return '—'
    return f'{secs // 60}:{secs % 60:02d}'


def _fmt_dt(dt, with_time=True):
    """Format a (possibly tz-aware) datetime/date for a spreadsheet cell."""
    if not dt:
        return ''
    if hasattr(dt, 'hour') and timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime('%Y-%m-%d %H:%M' if with_time else '%Y-%m-%d')


def _xlsx_response(filename, sheet_title, headers, rows, widths=None):
    """Build an .xlsx download from header labels + a list of row lists."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    head_fill = PatternFill('solid', fgColor='111827')
    head_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    wrap_top = Alignment(vertical='top', wrap_text=True)
    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = wrap_top

    for i, header in enumerate(headers, 1):
        col = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col].width = (widths or {}).get(header, 16)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def export_calls_xlsx(request):
    """Download the Call Log as an Excel file, honoring the current filters."""
    qs = _filtered_call_logs(request)
    headers = ['#', 'Number', 'Received By', 'SIM', 'Direction',
               'Status', 'Duration', 'Date & Time']
    rows = [[
        i, c.caller_number, c.received_by, c.sim, c.get_direction_display(),
        c.get_status_display(), _fmt_duration(c.duration), _fmt_dt(c.timestamp),
    ] for i, c in enumerate(qs, 1)]

    widths = {'Number': 18, 'Received By': 20, 'SIM': 16, 'Date & Time': 18}
    fname = f'call_logs_{timezone.now().date():%Y-%m-%d}.xlsx'
    return _xlsx_response(fname, 'Call Log', headers, rows, widths)


@login_required
def export_call_leads_xlsx(request):
    """Download the Call Leads list as an Excel file, honoring tab + filters."""
    qs = _filtered_call_leads(request)
    headers = ['#', 'Number', 'Call Status', 'Duration', 'Date & Time', 'Device',
               'SIM', 'Query', 'Quotation', 'Follow Up', 'Follow-up Note',
               'Notes', 'Lead Status', 'Returned Call']
    rows = [[
        i, l.caller_number, (l.call_status or '').title(), _fmt_duration(l.duration),
        _fmt_dt(l.call_time), l.received_by, l.sim, l.query, l.quotation_filename,
        _fmt_dt(l.follow_up, with_time=False), l.follow_up_note, l.notes,
        l.get_lead_status_display(), 'Yes' if l.return_called else '',
    ] for i, l in enumerate(qs, 1)]

    widths = {'Number': 18, 'Date & Time': 18, 'Device': 20, 'SIM': 16,
              'Query': 34, 'Quotation': 24, 'Follow-up Note': 28, 'Notes': 34}
    tab = request.GET.get('tab', 'active')
    fname = f'call_leads_{tab}_{timezone.now().date():%Y-%m-%d}.xlsx'
    return _xlsx_response(fname, 'Call Leads', headers, rows, widths)
